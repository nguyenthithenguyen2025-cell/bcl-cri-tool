# -*- coding: utf-8 -*-
"""Xuất dữ liệu BCL ra file Excel (openpyxl)."""

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.parameters import PARAMETERS, PARAM_BY_ID
from config.parameters import RISK_COLORS

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

RISK_LEVEL_LABELS = {
    1: "Cấp 1 — Rủi ro thấp",
    2: "Cấp 2 — Rủi ro trung bình",
    3: "Cấp 3 — Rủi ro cao",
    4: "Cấp 4 — Rủi ro rất cao",
}


def _header_cell(ws, row, col, value, bg="1f77b4", fg="FFFFFF", bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER
    return cell


def _data_cell(ws, row, col, value, bold=False, align="left", bg=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _sheet_bcl_info(wb, entries: list[dict]):
    """Sheet 1: Thông tin cơ bản BCL."""
    ws = wb.create_sheet("1. Thông tin BCL")
    ws.row_dimensions[1].height = 40

    headers = [
        "STT", "Tên BCL", "Tỉnh/TP", "Huyện", "Xã", "Loại BCL",
        "Diện tích (ha)", "Thể tích (m³)", "Chiều cao (m)",
        "Năm bắt đầu", "Năm ngừng", "Ghi chú",
    ]
    for col, h in enumerate(headers, 1):
        _header_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["L"].width = 40

    for i, e in enumerate(entries, 1):
        info = e.get("info", {})
        row_data = [
            i,
            info.get("ten_bcl", ""),
            info.get("tinh", ""),
            info.get("huyen", ""),
            info.get("xa", ""),
            info.get("loai_bcl", ""),
            info.get("dien_tich_ha"),
            info.get("the_tich_m3"),
            info.get("chieu_cao_m"),
            info.get("nam_bat_dau"),
            info.get("nam_ngung"),
            info.get("ghi_chu", ""),
        ]
        for col, val in enumerate(row_data, 1):
            _data_cell(ws, i + 1, col, val, align="center" if col == 1 else "left")


def _sheet_cri_scores(wb, entries: list[dict]):
    """Sheet 2: Điểm 14 thông số CRI."""
    ws = wb.create_sheet("2. Điểm CRI")
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 60

    # Tiêu đề cột: STT, Tên BCL, sau đó 14 thông số
    ws.cell(1, 1, "STT")
    ws.cell(1, 2, "Tên BCL")
    col = 3
    for p in PARAMETERS:
        _header_cell(ws, 1, col, f"{p['id']}\n{p['name']}", bg="2c3e50")
        ws.column_dimensions[get_column_letter(col)].width = 14
        col += 1
    _header_cell(ws, 1, col, "H")
    _header_cell(ws, 1, col + 1, "P")
    _header_cell(ws, 1, col + 2, "R")
    _header_cell(ws, 1, col + 3, "CRI", bg="e74c3c")
    _header_cell(ws, 1, col + 4, "Cấp rủi ro")
    _header_cell(ws, 1, col + 5, "Giải pháp KN")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28

    SCORE_COLORS = {
        0.25: "2ecc71", 0.50: "f39c12", 0.75: "e67e22", 1.00: "e74c3c"
    }

    for i, e in enumerate(entries, 1):
        r = i + 1
        scores = e.get("scores", {})
        result = e.get("result", {})
        info = e.get("info", {})

        _data_cell(ws, r, 1, i, align="center")
        _data_cell(ws, r, 2, info.get("ten_bcl", ""))

        col = 3
        for p in PARAMETERS:
            pid = p["id"]
            score = scores.get(pid)
            assumed = pid in result.get("assumed_max", [])
            val = score if score is not None else 1.00
            cell = _data_cell(ws, r, col, val, align="center",
                              bg=SCORE_COLORS.get(val))
            if assumed:
                cell.font = Font(italic=True, color="FFFFFF")
            col += 1

        _data_cell(ws, r, col, result.get("H"), align="center")
        _data_cell(ws, r, col + 1, result.get("P"), align="center")
        _data_cell(ws, r, col + 2, result.get("R"), align="center")

        cri = result.get("CRI")
        lv = result.get("risk", {}).get("level")
        lv_color = RISK_COLORS.get(lv, "#95a5a6").lstrip("#") if lv else "cccccc"

        cri_cell = _data_cell(ws, r, col + 3,
                              round(cri, 4) if cri else None,
                              bold=True, align="center", bg=lv_color)
        cri_cell.font = Font(bold=True, color="FFFFFF")
        _data_cell(ws, r, col + 4, RISK_LEVEL_LABELS.get(lv, "BCL-HVS"))
        _data_cell(ws, r, col + 5, result.get("solution", {}).get("short_name", ""))


def _sheet_results(wb, entries: list[dict]):
    """Sheet 3: Kết quả và giải pháp."""
    ws = wb.create_sheet("3. Kết quả & Giải pháp")

    _header_cell(ws, 1, 1, "Thông tin BCL và kết quả đánh giá CRI", bg="2c3e50")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 30

    FIELD_LABELS = [
        ("Tên BCL", "info.ten_bcl"),
        ("Tỉnh/TP", "info.tinh"),
        ("Loại BCL", "info.loai_bcl"),
        ("Diện tích (ha)", "info.dien_tich_ha"),
        ("Năm ngừng hoạt động", "info.nam_ngung"),
        ("Chỉ số H (Nguồn nguy hại)", "result.H"),
        ("Chỉ số P (Đường lan truyền)", "result.P"),
        ("Chỉ số R (Đối tượng tiếp nhận)", "result.R"),
        ("CRI tổng hợp", "result.CRI"),
        ("Cấp rủi ro", "result.risk.label"),
        ("Giải pháp khuyến nghị", "result.solution.name"),
        ("Hạng mục bắt buộc (tóm tắt)", "result.solution.mandatory_items"),
        ("Thời gian quản lý sau đóng", "result.solution.monitoring_period"),
        ("Căn cứ pháp lý", "result.solution.legal_basis"),
        ("Thông số thiếu dữ liệu", "missing_notes"),
    ]

    row = 2
    for entry in entries:
        info = entry.get("info", {})
        result = entry.get("result", {})
        risk = result.get("risk", {})
        solution = result.get("solution", {})

        for label, path in FIELD_LABELS:
            parts = path.split(".")
            val = None
            if parts[0] == "info":
                val = info.get(parts[1])
            elif parts[0] == "result":
                if len(parts) == 2:
                    val = result.get(parts[1])
                    if isinstance(val, float):
                        val = round(val, 4)
                elif parts[1] == "risk":
                    val = risk.get(parts[2])
                elif parts[1] == "solution":
                    val = solution.get(parts[2])
                    if isinstance(val, list):
                        val = "\n".join(f"• {x}" for x in val)
            elif parts[0] == "missing_notes":
                missing_notes = entry.get("missing_notes", {})
                val = (
                    "\n".join(f"{pid}: {note}" for pid, note in missing_notes.items() if note)
                    if missing_notes
                    else "—"
                )

            _header_cell(ws, row, 1, label, bg="ecf0f1", fg="2c3e50", bold=False)
            ws.cell(row, 2, value=val)
            ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 2).border = _BORDER
            ws.row_dimensions[row].height = 25
            row += 1

        # Dòng ngăn cách giữa các BCL
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 60


def export_to_excel(entries: list[dict]) -> BytesIO:
    """
    Xuất danh sách BCL ra file Excel với 3 sheet.

    Returns:
        BytesIO chứa file Excel.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # Xóa sheet mặc định

    _sheet_bcl_info(wb, entries)
    _sheet_cri_scores(wb, entries)
    _sheet_results(wb, entries)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
