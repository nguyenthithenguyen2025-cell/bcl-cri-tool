# -*- coding: utf-8 -*-
"""Trang 5 — So sánh nhiều BCL (bảng xếp hạng CRI)."""

import streamlit as st
import pandas as pd
from utils.session import get_bcl_summary_list, count_bcl
from utils.sidebar import render_sidebar
from utils.ui import apply_global_styles, render_page_footer, render_page_header

apply_global_styles()
render_sidebar()
render_page_header(
    "So sánh nhiều bãi chôn lấp",
    "Tổng hợp, lọc và xếp hạng các bãi chôn lấp theo kết quả CRI để hỗ trợ xác định ưu tiên "
    "xử lý hoặc lập kế hoạch đầu tư.",
    section="Dashboard so sánh",
)

if count_bcl() == 0:
    st.info(
        "💡 Chưa có BCL nào. Hãy vào trang **Khai báo BCL** → **Nhập thông số CRI** "
        "để thêm ít nhất một BCL. Thêm nhiều BCL để so sánh."
    )
    st.markdown("""
**Trang này sẽ hiển thị sau khi có dữ liệu:**
- Bảng xếp hạng tất cả BCL theo CRI giảm dần (màu nền theo cấp rủi ro)
- Bộ lọc theo tỉnh/thành, cấp rủi ro, giải pháp khuyến nghị
- Biểu đồ so sánh chỉ số H / P / R giữa các BCL
- Biểu đồ phân tán CRI theo diện tích BCL
- Xuất bảng xếp hạng sang Excel
    """)
    st.stop()

if count_bcl() == 1:
    st.info(
        "💡 Hiện chỉ có 1 BCL. Thêm BCL khác để so sánh. "
        "Dữ liệu một BCL vẫn được hiển thị dưới đây."
    )

summary = get_bcl_summary_list()


def priority_label(item: dict) -> str:
    """Xác định mức ưu tiên xử lý phục vụ lập kế hoạch."""
    if item.get("loai_bcl") == "HVS":
        return "Theo dõi vận hành"
    if item.get("CRI") is None:
        return "Cần hoàn thiện CRI"
    level = item.get("risk_level")
    if level == 4:
        return "Ưu tiên khẩn cấp"
    if level == 3:
        return "Ưu tiên cao"
    if level == 2:
        return "Ưu tiên trung bình"
    return "Ưu tiên thấp"


def build_rows(data: list[dict]) -> list[dict]:
    """Tạo dữ liệu bảng xếp hạng có định dạng trình bày."""
    rows = []
    for i, s in enumerate(data, 1):
        status_label = s.get("status_label", "BCL-HVS")
        cri_str = f"{s['CRI']:.4f}" if s["CRI"] is not None else status_label
        h_str = f"{s['H']:.4f}" if s["H"] is not None else "—"
        p_str = f"{s['P']:.4f}" if s["P"] is not None else "—"
        r_str = f"{s['R']:.4f}" if s["R"] is not None else "—"
        level = s["risk_level"]
        level_str = RISK_LEVEL_LABELS.get(level, status_label)
        rows.append({
            "STT": i,
            "Tên BCL": s["ten_bcl"],
            "Tỉnh": s["tinh"],
            "DT (ha)": s["dien_tich_ha"],
            "H": h_str,
            "P": p_str,
            "R": r_str,
            "CRI": cri_str,
            "Cấp rủi ro": level_str,
            "Ưu tiên xử lý": priority_label(s),
            "Giải pháp KN": s["solution_name"],
            "ID": s["id"],
            "_risk_level": level,
        })
    return rows


def portfolio_metrics(data: list[dict]) -> dict:
    total = len(data)
    khvs_done = sum(1 for item in data if item.get("CRI") is not None)
    high = sum(1 for item in data if item.get("risk_level") == 3)
    very_high = sum(1 for item in data if item.get("risk_level") == 4)
    urgent = sum(1 for item in data if priority_label(item) in ("Ưu tiên cao", "Ưu tiên khẩn cấp"))
    pending = sum(1 for item in data if item.get("loai_bcl") != "HVS" and item.get("CRI") is None)
    return {
        "total": total,
        "khvs_done": khvs_done,
        "high": high,
        "very_high": very_high,
        "urgent": urgent,
        "pending": pending,
    }


st.subheader("Tổng quan danh mục BCL")
metrics = portfolio_metrics(summary)
metric_cols = st.columns(5)
metric_items = [
    ("Tổng số BCL", metrics["total"]),
    ("Đã tính CRI", metrics["khvs_done"]),
    ("Rủi ro cao", metrics["high"]),
    ("Rủi ro rất cao", metrics["very_high"]),
    ("Cần ưu tiên", metrics["urgent"]),
]
for col, (label, value) in zip(metric_cols, metric_items):
    with col:
        st.metric(label, value)

if metrics["pending"] > 0:
    st.warning(f"Có {metrics['pending']} BCL-KHVS chưa tính CRI; cần hoàn thiện trước khi dùng bảng ưu tiên xử lý.")

st.divider()

# ── Bộ lọc
st.subheader("Bộ lọc")
filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4)

all_tinh = sorted(set(s["tinh"] for s in summary if s["tinh"]))
all_levels = sorted(set(s["risk_level"] for s in summary if s["risk_level"] is not None))
all_solutions = sorted(set(s["solution_name"] for s in summary if s["solution_name"]))
all_priorities = sorted(set(priority_label(s) for s in summary))

with filter_col1:
    sel_tinh = st.multiselect("Lọc theo tỉnh/thành:", options=all_tinh)
with filter_col2:
    sel_levels = st.multiselect(
        "Lọc theo cấp rủi ro:",
        options=all_levels,
        format_func=lambda x: f"Cấp {x}",
    )
with filter_col3:
    sel_solutions = st.multiselect("Lọc theo giải pháp:", options=all_solutions)
with filter_col4:
    sel_priorities = st.multiselect("Lọc theo ưu tiên:", options=all_priorities)

# Áp dụng bộ lọc
filtered = summary
if sel_tinh:
    filtered = [s for s in filtered if s["tinh"] in sel_tinh]
if sel_levels:
    filtered = [s for s in filtered if s["risk_level"] in sel_levels]
if sel_solutions:
    filtered = [s for s in filtered if s["solution_name"] in sel_solutions]
if sel_priorities:
    filtered = [s for s in filtered if priority_label(s) in sel_priorities]

st.caption(f"Hiển thị {len(filtered)}/{len(summary)} BCL")
st.divider()

# ════════════════════════════════════════════════════════
# BẢNG XẾP HẠNG
# ════════════════════════════════════════════════════════
st.subheader("Bảng xếp hạng theo CRI")

RISK_LEVEL_LABELS = {1: "Cấp 1 — Thấp", 2: "Cấp 2 — TB", 3: "Cấp 3 — Cao", 4: "Cấp 4 — Rất cao"}
rows = build_rows(filtered)

df = pd.DataFrame(rows)
if not df.empty:
    display_df = df.drop(columns=["ID", "_risk_level"])
    st.dataframe(
        display_df,
        use_container_width=True,
        height=min(400, 60 + 40 * len(df)),
        hide_index=True,
    )
else:
    st.info("Không có BCL nào khớp với bộ lọc.")

st.divider()

# ════════════════════════════════════════════════════════
# BIỂU ĐỒ SO SÁNH
# ════════════════════════════════════════════════════════
if len(filtered) > 0:
    from utils.charts import (
        comparison_bar_chart,
        priority_top_chart,
        province_distribution_chart,
        risk_distribution_chart,
        scatter_cri_area,
    )

    st.subheader("Biểu đồ so sánh")

    khvs_data = [s for s in filtered if s["CRI"] is not None]

    col_d1, col_d2 = st.columns(2)
    with col_d1:
        st.markdown("**Phân bố số BCL theo cấp rủi ro**")
        fig_risk = risk_distribution_chart(filtered)
        if fig_risk:
            st.plotly_chart(fig_risk, use_container_width=True)
    with col_d2:
        st.markdown("**Phân bố số BCL theo tỉnh/thành phố**")
        fig_province = province_distribution_chart(filtered)
        if fig_province:
            st.plotly_chart(fig_province, use_container_width=True)

    if khvs_data:
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            st.markdown("**So sánh chỉ số H / P / R của các BCL-KHVS**")
            fig_cbar = comparison_bar_chart(khvs_data)
            if fig_cbar:
                st.plotly_chart(fig_cbar, use_container_width=True)

        with col_c2:
            st.markdown("**CRI theo diện tích BCL**")
            fig_scatter = scatter_cri_area(khvs_data)
            if fig_scatter:
                st.plotly_chart(fig_scatter, use_container_width=True)

        st.markdown("**Top BCL-KHVS cần ưu tiên theo CRI**")
        fig_priority = priority_top_chart(khvs_data, n=10)
        if fig_priority:
            st.plotly_chart(fig_priority, use_container_width=True)
    else:
        st.info("Không có BCL-KHVS nào trong bộ lọc hiện tại để hiển thị biểu đồ.")

    st.divider()

# ════════════════════════════════════════════════════════
# XUẤT DANH SÁCH EXCEL
# ════════════════════════════════════════════════════════
st.subheader("Xuất danh sách")

if not df.empty:
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def make_ranking_excel(data: list[dict]) -> BytesIO:
        wb = openpyxl.Workbook()
        ws0 = wb.active
        ws0.title = "Tổng quan"

        metrics_all = portfolio_metrics(data)
        overview_rows = [
            ("Tổng số BCL", metrics_all["total"]),
            ("BCL-KHVS đã tính CRI", metrics_all["khvs_done"]),
            ("BCL rủi ro cao", metrics_all["high"]),
            ("BCL rủi ro rất cao", metrics_all["very_high"]),
            ("BCL cần ưu tiên xử lý", metrics_all["urgent"]),
            ("BCL-KHVS chưa tính CRI", metrics_all["pending"]),
        ]
        ws0.append(["Chỉ tiêu", "Giá trị"])
        for cell in ws0[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1f77b4")
        for row in overview_rows:
            ws0.append(row)
        ws0.column_dimensions["A"].width = 34
        ws0.column_dimensions["B"].width = 16

        ws = wb.create_sheet("Bảng ưu tiên BCL")

        headers = [
            "STT", "Tên BCL", "Tỉnh/TP", "Diện tích (ha)",
            "H", "P", "R", "CRI", "Cấp rủi ro", "Ưu tiên xử lý", "Giải pháp khuyến nghị",
        ]
        ws.append(headers)

        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1f77b4")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        LEVEL_COLORS = {1: "2ecc71", 2: "f39c12", 3: "e67e22", 4: "e74c3c"}

        for i, s in enumerate(data, 1):
            row = [
                i,
                s["ten_bcl"],
                s["tinh"],
                s["dien_tich_ha"],
                f"{s['H']:.4f}" if s["H"] else "—",
                f"{s['P']:.4f}" if s["P"] else "—",
                f"{s['R']:.4f}" if s["R"] else "—",
                f"{s['CRI']:.4f}" if s["CRI"] is not None else s.get("status_label", "BCL-HVS"),
                RISK_LEVEL_LABELS.get(s["risk_level"], s.get("status_label", "BCL-HVS")),
                priority_label(s),
                s["solution_name"],
            ]
            ws.append(row)

            level = s["risk_level"]
            if level and level in LEVEL_COLORS:
                fill = PatternFill("solid", fgColor=LEVEL_COLORS[level] + "40")
                for cell in ws[ws.max_row]:
                    cell.fill = fill

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        for col in ["D", "E", "F", "G", "H"]:
            ws.column_dimensions[col].width = 12
        ws.column_dimensions["I"].width = 22
        ws.column_dimensions["J"].width = 22
        ws.column_dimensions["K"].width = 20

        ws_stats = wb.create_sheet("Thống kê theo tỉnh")
        ws_stats.append(["Tỉnh/TP", "Tổng BCL", "Đã tính CRI", "Rủi ro cao", "Rủi ro rất cao", "Cần ưu tiên"])
        for cell in ws_stats[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1f77b4")

        by_province: dict[str, list[dict]] = {}
        for item in data:
            province = item.get("tinh") or "Chưa rõ địa phương"
            by_province.setdefault(province, []).append(item)

        for province, items in sorted(by_province.items()):
            m = portfolio_metrics(items)
            ws_stats.append([province, m["total"], m["khvs_done"], m["high"], m["very_high"], m["urgent"]])
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_stats.column_dimensions[col].width = 20

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    excel_buf = make_ranking_excel(filtered)
    st.download_button(
        label="📥 Tải xuống bảng xếp hạng (.xlsx)",
        data=excel_buf,
        file_name="Bang_xep_hang_BCL_CRI.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

render_page_footer()
